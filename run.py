import json
import datetime
from random import uniform

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from transliterate import translit

from tools import fix_width, create_cell

wb = load_workbook(filename='base.xlsx')
ws = wb.worksheets[1]

with open('jobs.json', encoding='utf-8') as data_file:
    data = json.load(data_file)
jobs_list = data['jobs']

person_name = data['person_name']
create_cell(ws['A3'], value=person_name)
create_cell(ws['A4'], value=data['person_position'])

work_hours = data['work_hours']
work_minutes = work_hours * 60
create_cell(ws['A5'], value='График работы %s часов' % str(work_hours))
create_cell(ws['B5'], value=work_minutes)

header_row = data['header_row']
lines_delta = data['lines_delta']

weekends_background = PatternFill("solid", fgColor="dddddd")
border_style = Side(border_style="dashed", color="000000")
border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
left_alignment = Alignment(horizontal="left", vertical="center")
right_alignment = Alignment(horizontal="right", vertical="center")

header_cell = ws['A' + str(header_row)]
create_cell(header_cell, border=border, alignment=left_alignment)

row_number = header_row
for job in jobs_list:
    row_number += 1
    create_cell(ws['A' + str(row_number)], value=job['name'], border=border, alignment=left_alignment)

create_cell(
    ws['A' + str(row_number + 1 + lines_delta)],
    value='Трудозатраты в минутах',
    border=border,
    alignment=left_alignment
)

create_cell(
    ws['A' + str(row_number + 2 + lines_delta)],
    value='Трудозатраты в часах',
    border=border,
    alignment=left_alignment
)

chars_list = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q',
              'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF']

today = datetime.date.today()
current_month = today.month
current_year = today.year

for char in chars_list:
    row_number = header_row
    day = chars_list.index(char) + 1
    try:
        cell_date = datetime.date(year=current_year, month=current_month, day=day)
    except ValueError:
        cell_date = None

    if cell_date:
        if cell_date.weekday() < 5:
            cell_value = cell_date
            cell_fill = None
        else:
            cell_value = '###'
            cell_fill = weekends_background
        create_cell(
            ws[char + str(row_number)],
            value=cell_value,
            border=border,
            alignment=right_alignment,
            fill=cell_fill
        )

        total_duration = 0
        for job in jobs_list:
            row_number += 1
            if cell_date.weekday() < 5:
                job_duration = job['duration']
                random_delta = uniform(-job_duration / 6, job_duration / 6)
                final_duration = job_duration + round(random_delta)
                job_cell_value = final_duration
                total_duration += final_duration
                job_cell_fill = None
            else:
                job_cell_value = None
                job_cell_fill = weekends_background
            create_cell(
                ws[char + str(row_number)],
                border=border,
                alignment=right_alignment,
                value=job_cell_value,
                fill=job_cell_fill
            )

        for i in range(1, lines_delta + 1):
            create_cell(ws[char + str(row_number + i)], border=border)

        create_cell(
            ws[char + str(row_number + 1 + lines_delta)],
            border=border,
            alignment=right_alignment,
            value=total_duration
        )

        create_cell(
            ws[char + str(row_number + 2 + lines_delta)],
            border=border,
            alignment=right_alignment,
            value=round(total_duration / 60, 1)
        )

        create_cell(
            ws[char + str(row_number + 3 + lines_delta)],
            alignment=right_alignment,
            value='%s%%' % int((round(total_duration / work_minutes, 2) * 100))
        )

fix_width(ws)

file_name = '%s_%s.xlsx' % (translit(person_name, 'ru', reversed=True).replace(' ', '_'), today.strftime('%B'))
wb.save(file_name)
